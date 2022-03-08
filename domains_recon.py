from openpyxl import load_workbook
import requests



filename = './example.xlsx'

def check_attributes():
    workbook = load_workbook(filename=filename)

    new_cols = ['Web Application URL', 'Publicly / Internally Accessible', 'URL Redirected to',
                'Active / Redirect / Parked']

    # Write all the new column headers for all sheets
    sheets = workbook.sheetnames
    for sheet in sheets:
        current_sheet = workbook[sheet]
        number_of_cols = len([x for x in current_sheet[1] if x.value])
        i = number_of_cols
        for item in new_cols:
            new_col_coords = chr(65 + i) + '1'
            current_sheet[new_col_coords] = item
            i += 1

        # For each sheet check if URL is set up with htpps, http, or neither and write to column Web Application URL
        k = 1
        for domain in current_sheet['A']:
            if k == 1:
                k += 1
                continue

            url = check_https(domain.value)
            url_cell_coord = chr(65 + number_of_cols) + str(k)

            current_sheet[url_cell_coord] = url

            publicly_avail_coord = chr(65 + number_of_cols + 1) + str(k)
            if url:
                current_sheet[publicly_avail_coord] = 'yes'
            else:
                current_sheet[publicly_avail_coord] = 'no'

            url_redirect = check_redirect(domain.value)
            url_redirect_coord = chr(65 + number_of_cols + 2) + str(k)
            current_sheet[url_redirect_coord] = url_redirect

            domain_status_coord = chr(65 + number_of_cols + 3) + str(k)
            if url_redirect:
                current_sheet[domain_status_coord] = 'redirect'
            elif url and not url_redirect:
                current_sheet[domain_status_coord] = 'active'
            else:
                current_sheet[domain_status_coord] = 'Parked'
            k += 1

    workbook.save(filename=filename)


def check_https(domain):
    '''Checks to see if domain has a https or http url'''
    try:
        # Check to see if URL is https enabled
        url = f'https://{domain}'
        response = requests.get(url, timeout=1)
        return url
    # if not check http
    except requests.exceptions.RequestException as e:
        try:
            url = f'http://:{domain}'
            response = requests.get(url, timeout=1)
            return url
        # If there is an error it means no connection via https or http
        except requests.exceptions.RequestException as e:
            return ''


def check_redirect(domain):
    '''Checks to see if it redirects, 
    wil not give you full history of redirects only the final redirect.'''
    try:
        responses = requests.get('http://' + domain, timeout=1)
        if len(responses.history) > 1:

            return responses.history[-1].headers['location']

    except requests.exceptions.RequestException as e:
        try:
            responses = requests.get('http://www.' + domain, timeout=1)
            if len(responses.history) > 1:
                return responses.history[-1].headers['location']

        except requests.exceptions.RequestException as e:
            return ''



check_attributes()
